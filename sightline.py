#!/usr/bin/env python
"""AGOL web-map sharing audit: which maps contain layers their viewers cannot see.

Answers one question per layer reference: is every audience member of the MAP also an
audience member of the LAYER?  Audience containment, not an access "rank" -- 'org' and
'shared' are incomparable audiences, not equal ones (an org member outside the group
genuinely cannot load a group-shared layer).

HARD RULE -- never attribute-access an unknown field on an Item.  Item.__getattr__ does a
network round trip *before* raising AttributeError (measured 0.665s each); the original
script spent ~95% of its runtime inside two such probes for columns that were empty in
171/171 rows.  Dict access only: it["access"], it.get("numViews", 0).  The only permitted
attribute/method calls on an Item are .get_data() and .id.

Run:  python sightline.py [-o PATH] [--limit N] [--deep] [--self-check]
"""

import os
import re
import sys
import json
import time
import getpass
import argparse
import datetime
import urllib.error
import urllib.parse
import urllib.request
import collections

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from arcgis.gis import GIS

# =============================================================================
#  CONFIGURATION -- everything you may want to change lives in this block.
#  Every value here is a default; the matching command-line flag wins when given.
# =============================================================================

# --- connection ---------------------------------------------------------------
# Portal URL.  ArcGIS Online for everyone except ArcGIS Enterprise sites, which
# use e.g. "https://gis.yourcounty.org/portal".
PORTAL_URL = "https://www.arcgis.com"

# Leave USERNAME blank to use, in order: $ARCGIS_USER -> the signed-in ArcGIS Pro
# session -> an interactive prompt.
USERNAME = ""

# PASSWORD: leave blank.  Set $ARCGIS_PASS instead, or let the script prompt.
# A password typed here is a password you will eventually commit to git.
PASSWORD = ""

# --- what to scan -------------------------------------------------------------
MAX_MAPS = 10000        # ceiling on web maps fetched
DEFAULT_LIMIT = None    # None = scan everything; set an int to smoke-test (--limit)
DEEP_DEFAULT = False    # True = always collect FeatureCount + LastEditUTC (--deep)

# --- output -------------------------------------------------------------------
OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
OUTPUT_PREFIX = "sightline"              # final name: <prefix>_YYYYmmdd_HHMMSS.xlsx
MAX_COL_WIDTH = 55                  # Excel column width cap
WIDTH_SAMPLE_ROWS = 400             # rows sampled when sizing columns

# --- service probing ----------------------------------------------------------
# The probe asks, anonymously, "can the public actually load this service?"
PROBE_TIMEOUT = 8       # seconds per service URL
MAX_HOST_FAILS = 3      # consecutive transport failures before a host is skipped.
                        # Without this, one unreachable on-prem box x 50 urls x 8s
                        # is a 400s hang inside a ~160s run.
UA = {"User-Agent": "Mozilla/5.0"}

# TLS verification stays ON.  A service whose certificate does not verify is one an
# anonymous browser would also refuse -- that is a finding (UNREACHABLE /
# SSLCertVerificationError), not something to paper over with an unverified context.

# ArcGIS REST error codes.  400 is not a typo: a request for a service that no longer
# exists answers "Invalid URL" with code 400, not 404 (verified against 12 dead urls --
# an org_admin session gets the identical 400, so it is absence, not permissions).
SECURED_CODES = {401, 403, 498, 499}
DEAD_CODES = {400, 404}

# --- batching (see README "Why it is fast") -----------------------------------
ITEM_CHUNK = 50         # ids per id:(a OR b ...) search; 18.8x faster than one-by-one
GROUP_CHUNK = 80        # ids per content/itemsgroups call

# --- report shape -------------------------------------------------------------
# Severity is driven by the MAP's audience: a broken layer in a public map matters
# more than the same layer in a private one.  Sorts correctly as text.
SEV = {"public": "1 High", "org": "2 Medium", "shared": "3 Low", "private": "4 Info"}

FIND_COLS = ["Severity", "Verdict", "Reason", "MapTitle", "MapOwner", "MapAccess",
             "MapViews", "MapUrl", "LayerTitle", "LayerAccess", "LayerUrl"]
ALL_COLS = ["MapTitle", "MapId", "MapOwner", "MapAccess", "MapViews", "MapGroups", "MapUrl",
            "Section", "LayerTitle", "LayerType", "LayerItemId", "LayerUrl", "LayerHost",
            "LayerOwner", "LayerAccess", "LayerGroups", "LayerState", "Verdict", "Severity",
            "Reason", "Overexposed"]

# =============================================================================
#  END CONFIGURATION
# =============================================================================

FS_RE = re.compile(r"/featureserver/\d+/?$")          # strict: the service ROOT cannot be counted


# --------------------------------------------------------------------------- plumbing

def connect():
    """Resolve credentials, most explicit first, and sign in.

    CONFIG block -> environment -> signed-in ArcGIS Pro session -> interactive prompt.
    A password is never written to disk by this script.
    """
    # PORTAL_URL always holds a value, so $ARCGIS_URL overrides it; USERNAME and
    # PASSWORD default to blank, so a value typed in CONFIG overrides the environment.
    url = os.environ.get("ARCGIS_URL") or PORTAL_URL
    user = USERNAME or os.environ.get("ARCGIS_USER")
    pw = PASSWORD or os.environ.get("ARCGIS_PASS")
    if user and pw:
        return GIS(url, user, pw)
    if not user:
        try:
            return GIS("pro")
        except Exception as e:
            print("No active ArcGIS Pro session (%s); falling back to prompt." % type(e).__name__)
        user = input("Username: ")
    return GIS(url, user, pw or getpass.getpass("Password: "))


def iso(ms):
    """epoch MILLISECONDS -> 'YYYY-MM-DD HH:MM:SS+00:00' UTC.  '' when falsy."""
    if not ms:
        return ""
    dt = datetime.datetime.fromtimestamp(ms / 1000.0, datetime.timezone.utc)
    return str(dt.replace(microsecond=0))


def chunks(seq, n):
    seq = list(seq)
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def host(url):
    try:
        return urllib.parse.urlparse(url or "").netloc
    except ValueError:
        return ""


def walk(data, out):
    """Append (section, layer_dict) for every real layer reference in a web map's JSON."""
    if not isinstance(data, dict):
        return
    for section in ("operationalLayers", "tables"):
        for lyr in (data.get(section) or []):
            _emit(section, lyr, out)
    for lyr in ((data.get("baseMap") or {}).get("baseMapLayers") or []):
        _emit("baseMap", lyr, out)


def _emit(section, lyr, out):
    if not isinstance(lyr, dict):
        return
    # ponytail: children are filtered to entries with a target, so MapServer sublayer
    # style-overrides (which also live under "layers") never leak in as phantom rows.
    kids = [c for c in (lyr.get("layers") or [])
            if isinstance(c, dict) and (c.get("url") or c.get("itemId"))]
    if lyr.get("url") or lyr.get("itemId") or not kids:
        out.append((section, lyr))          # a GroupLayer shell with children is not itself a ref
    for c in kids:
        _emit(section, c, out)


# --------------------------------------------------------------------------- phase 1-3

def collect(gis, maps):
    """One get_data per map (~0.09s).  A failed map yields an error tuple, never vanishes."""
    base = gis.url.rstrip("/") + "/home/item.html?id="
    refs, errors, t0, n = [], [], time.time(), len(maps)
    for i, m in enumerate(maps, 1):
        meta = {"MapTitle": m["title"], "MapId": m["id"], "MapOwner": m["owner"],
                "MapAccess": m["access"], "MapViews": m.get("numViews", 0),
                "MapUrl": base + m["id"]}
        found = []
        try:
            # walk() must stay inside the try: nested-corrupt map JSON (a string where
            # baseMap should be, an int under "tables") raises there, not in get_data,
            # and one such map would otherwise abort the whole run with no error row.
            walk(m.get_data(try_json=True), found)
        except Exception as e:
            errors.append((meta, "%s: %s" % (type(e).__name__, e)))
            continue
        for section, lyr in found:
            refs.append((meta, section, lyr))
        if i % 25 == 0 or i == n:
            el = time.time() - t0
            print("  %d/%d maps  %.0fs elapsed  ~%.0fs left" % (i, n, el, el / i * (n - i)))
    return refs, errors


def resolve_items(gis, ids):
    """Batched id lookup -> ({id: Item}, {ids whose lookup never completed}).

    18.8x faster than gis.content.get() one at a time, same results.  The second return
    value matters: an id missing because its chunk errored is NOT the same as an id
    missing because the item was deleted, and reporting a throttled request as a deleted
    layer would put up to ITEM_CHUNK false DEAD rows in the report.
    """
    out, unresolved = {}, set()
    for chunk in chunks(sorted(ids), ITEM_CHUNK):
        try:
            hits = gis.content.search(query="id:(" + " OR ".join(chunk) + ")",
                                      max_items=len(chunk) * 2, outside_org=True)
        except Exception as e:
            print("  item batch failed (%s: %s) -- %d ids unresolved" % (type(e).__name__, e, len(chunk)))
            unresolved.update(chunk)
            continue
        for it in hits:
            out[it.id] = it
    return out, unresolved


def fetch_groups(gis, ids):
    """itemId -> {groupId} and groupId -> title, 80 ids per round trip.

    Ids from a FAILED chunk stay absent so .get() returns None (UNKNOWN); ids from a
    successful chunk are pre-seeded with an empty set (genuinely shared to no group).
    """
    gids, titles = {}, {}
    url = gis._portal.resturl + "content/itemsgroups"
    for chunk in chunks(sorted(ids), GROUP_CHUNK):
        try:
            resp = gis._con.get(url, {"f": "json", "items": ",".join(chunk)}) or {}
        except Exception as e:
            print("  group batch failed (%s: %s) -- %d ids unknown" % (type(e).__name__, e, len(chunk)))
            continue
        for iid in chunk:
            gids.setdefault(iid, set())
        for gid, blk in resp.items():
            blk = blk or {}
            titles[gid] = blk.get("groupTitle") or gid
            for ent in (blk.get("items") or []):
                iid = (ent or {}).get("itemId")
                if iid:
                    gids.setdefault(iid, set()).add(gid)
    return gids, titles


# --------------------------------------------------------------------------- phase 4 probe

def code_state(code):
    if code in SECURED_CODES:
        return ("SECURED", "code %s" % code)
    # ponytail: ArcGIS REST answers a request for a service that no longer exists with
    # 400 "Invalid URL", not 404.  Verified against 12 distinct dead urls: the org_admin
    # session gets the identical 400, so it is absence, not a permission problem.
    if code in DEAD_CODES:
        return ("DEAD", "code %s" % code)
    return ("UNKNOWN", "code %s" % code)


def probe(url, cache, fails):
    """One anonymous GET url?f=json.  Answers: can the public actually load this service?"""
    key = (url or "").split("?")[0]
    if key in cache:
        return cache[key]
    h = host(key)
    # ponytail: per-host circuit breaker. Without it one unreachable on-prem box x 50 urls
    # x 8s timeout is a 400s hang. Reset on any completed round trip -- a 499 means alive.
    if fails[h] >= MAX_HOST_FAILS:
        return ("UNREACHABLE", "host circuit-open")
    try:
        with urllib.request.urlopen(urllib.request.Request(key + "?f=json", headers=UA),
                                    timeout=PROBE_TIMEOUT) as r:
            raw = r.read()
    except urllib.error.HTTPError as e:
        fails[h] = 0
        res = code_state(e.code)
    except Exception as e:
        fails[h] += 1
        res = ("UNREACHABLE", type(e).__name__)
    else:
        fails[h] = 0
        try:
            body = json.loads(raw.decode("utf-8", "replace"))
        except ValueError:
            body = None
        if not isinstance(body, dict):
            res = ("UNKNOWN", "non-JSON response")
        else:
            err = body.get("error") or {}
            res = code_state(err.get("code", 0)) if err else ("PUBLIC", "")
    cache[key] = res
    return res


# --------------------------------------------------------------------------- phase 5 verdict

def covers(la, lg, lown, ma, mg, mown, titles=None):
    """Is every viewer of the map also a viewer of the layer?  (True / False / None=unknown)"""
    if la == "public":
        return True, "public layer covers every audience"
    if ma == "private":
        if la == "private" and lown != mown:
            return False, "private layer owned by %s" % (lown or "another user")
        return True, "private map; the audience is the owner only"
    if la == "org":
        if ma == "public":
            return False, "org-only layer in a public map; anonymous viewers cannot load it"
        return True, "org layer covers org and group audiences"
    if la == "shared":
        if ma != "shared":
            return False, "%s map, group-restricted layer" % ma
        # Both sides must be known.  Treating an unavailable MAP group set as "shared to no
        # groups" would silently score a never-performed check as OK -- the one failure mode
        # this tool must not have.  A failed fetch_groups chunk is exactly how that happens.
        if lg is None or mg is None:
            return None, "group membership unavailable"
        missing = set(mg) - lg
        if missing:
            names = sorted((titles or {}).get(g, g) for g in missing)
            return False, "map shared to groups not on the layer: " + ", ".join(names)
        return True, "layer shared to every group the map is shared to"
    return False, "private layer in a %s map" % ma


def verdict(state, detail, la, lg, lown, ma, mg, mown, titles=None):
    if ma == "shared" and mg is not None and not mg:
        ma = "private"                      # shared to zero groups == owner-only audience
    if state == "MISSING_ITEM":
        return "DEAD", "layer item deleted or inaccessible, and no service url to probe"
    if state == "DEAD":
        return "DEAD", "service gone (%s)" % (detail or "no longer published")
    if state == "EMBEDDED":
        return "OK", "feature collection embedded in the map; nothing external to share"
    if state in ("UNREACHABLE", "UNKNOWN", "NO_TARGET"):
        return "UNKNOWN", detail or "layer reference has no item id and no service url"
    if state == "SECURED":
        if ma == "public":
            return "BROKEN", "secured service in a public map; anonymous viewers cannot load it"
        return "UNKNOWN", "secured service; on-prem credentials are not verifiable from AGOL"
    ok, why = covers(la, lg, lown, ma, mg, mown, titles)
    return ("UNKNOWN" if ok is None else "OK" if ok else "BROKEN"), why


def build(refs, items, gids, gtitle, probes, fails, org_id, unresolved=()):
    rows = []
    for meta, section, lyr in refs:
        iid, url = lyr.get("itemId") or "", lyr.get("url") or ""
        it = items.get(iid) if iid else None
        la = lown = None
        lg, detail = None, ""
        if it is not None:
            state, la, lown, lg = "item", it["access"], it["owner"], gids.get(it.id)
        elif url:
            # Probe the SERVICE, not just the item: a deleted item whose service is still
            # public means the layer still draws.  Short-circuiting on the batch miss
            # would call that DEAD -- a whole false-positive class, removed by one branch.
            state, detail = probe(url, probes, fails)
            if state == "PUBLIC":
                la, lown, lg = "public", "", set()
        elif iid in unresolved:
            # its lookup errored, so we never learned anything -- do not claim it is deleted
            state, detail = "UNKNOWN", "item lookup failed; membership not verified"
        elif iid:
            state = "MISSING_ITEM"
        elif lyr.get("featureCollection"):
            state = "EMBEDDED"
        else:
            state = "NO_TARGET"
        mg = gids.get(meta["MapId"])
        v, why = verdict(state, detail, la, lg, lown, meta["MapAccess"], mg,
                         meta["MapOwner"], gtitle)
        # Overexposed is a DIFFERENT failure mode on a different axis; never in Findings.
        # The org-owned gate suppresses public Esri basemap noise.
        over = "Yes" if (la == "public" and meta["MapAccess"] != "public"
                         and it is not None and it.get("orgId") == org_id) else ""
        row = dict(meta)
        row.update({
            "Section": section, "LayerTitle": lyr.get("title") or "", "LayerType": lyr.get("layerType") or "",
            "LayerItemId": iid, "LayerUrl": url, "LayerHost": host(url), "LayerOwner": lown or "",
            "LayerAccess": la or "", "LayerState": state, "Verdict": v, "Reason": why,
            "Severity": "" if v == "OK" else SEV.get(meta["MapAccess"], "4 Info"),
            "Overexposed": over,
            "MapGroups": "; ".join(sorted(gtitle.get(g, g) for g in (mg or ()))),
            "LayerGroups": "; ".join(sorted(gtitle.get(g, g) for g in (lg or ()))),
        })
        rows.append(row)
    return rows


def error_rows(errors):
    """An error IS a finding -- it lands in Findings automatically, so there is no Errors sheet."""
    out = []
    for meta, msg in errors:
        row = dict(meta)
        row.update({"Section": "<error>", "Verdict": "UNKNOWN", "Reason": msg,
                    "Severity": SEV.get(meta["MapAccess"], "4 Info"), "LayerState": "UNKNOWN"})
        out.append(row)
    return out


# --------------------------------------------------------------------------- phase 6-7

def deep_stats(gis, urls):
    """--deep only.  Cached BY URL: many web-map layers share one FeatureServer endpoint."""
    from arcgis.features import FeatureLayer
    out = {}
    for u in sorted(urls):
        if not FS_RE.search(u.lower()):
            continue
        try:
            fl = FeatureLayer(u, gis)
            out[u] = (fl.query(return_count_only=True),
                      iso((fl.properties.get("editingInfo") or {}).get("lastEditDate")))
        except Exception as e:
            out[u] = ("", type(e).__name__)
    return out


def summarize(rows, meta):
    # columns= is load-bearing: an org with no web maps yields rows == [], and a bare
    # DataFrame([]) has no columns at all, so df.Verdict raises AttributeError.
    df = pd.DataFrame(rows, columns=ALL_COLS)
    out = [("Run", k, v) for k, v in meta.items()]
    for v in ("OK", "BROKEN", "DEAD", "UNKNOWN"):
        out.append(("Verdicts", v, int((df.Verdict == v).sum())))
    out.append(("Verdicts", "Overexposed (public layer, non-public map)",
                int((df.Overexposed == "Yes").sum()) if "Overexposed" in df else 0))
    bad = df[df.Verdict != "OK"]
    worst = bad.groupby(["MapTitle", "MapAccess", "MapViews"]).size().sort_values(ascending=False)
    for (t, a, vw), n in list(worst.items())[:15]:
        out.append(("Worst maps", "%s (%s, %s views)" % (t, a, vw), int(n)))
    for owner, n in list(bad.groupby("MapOwner").size().sort_values(ascending=False).items())[:10]:
        out.append(("Top owners", owner, int(n)))          # this is who you email
    return pd.DataFrame(out, columns=["Section", "Metric", "Value"])


def save_report(path, findings, all_rows, summary):
    """write_excel, but never lose a completed audit to a locked output file.

    Re-running with the previous report still open in Excel is the single most likely
    real-world failure for this audience, and it would otherwise discard a run that has
    already spent minutes of API calls.
    """
    try:
        write_excel(path, findings, all_rows, summary)
        return path
    except PermissionError:
        stem, ext = os.path.splitext(path)
        alt = "%s_%s%s" % (stem, datetime.datetime.now().strftime("%H%M%S"), ext)
        print("  %s is locked (open in Excel?); writing %s instead" % (path, alt))
        write_excel(alt, findings, all_rows, summary)
        return alt


def defuse(ws):
    """Stop a hostile item title from becoming a live formula in the report.

    A layer or map title is attacker-controllable -- any org member can name an item
    `=cmd|' /C calc'!A0`, and openpyxl faithfully types a leading =/+/-/@ as a formula.
    Forcing the cell's type to string keeps the text exactly as-is (an apostrophe prefix
    would corrupt the value) while making it inert.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cell.data_type = "s"


def scrub(df):
    """Drop control characters openpyxl refuses to write.

    Unreachable through the AGOL web UI, reachable through the REST API, and it would
    otherwise raise IllegalCharacterError at the very end of a multi-minute run.
    """
    # ponytail: guard on the value, not the column dtype -- pandas 3 types text columns as
    # StringDtype rather than object, so a `dtype == object` test silently scrubs nothing.
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(
            lambda v: ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v)
    return out


def write_excel(path, findings, all_rows, summary):
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, df, links in (("Findings", findings, ("MapUrl", "LayerUrl")),
                                ("AllLayers", all_rows, ("MapUrl", "LayerUrl")),
                                ("Summary", summary, ())):
            df = scrub(df)
            df.to_excel(xw, index=False, sheet_name=name)
            ws = xw.sheets[name]
            defuse(ws)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i, col in enumerate(df.columns, 1):
                widths = [len(str(col))] + [len(str(v)) for v in df[col].head(WIDTH_SAMPLE_ROWS)]
                ws.column_dimensions[get_column_letter(i)].width = min(max(widths) + 2, MAX_COL_WIDTH)
                if col not in links:
                    continue
                for r in range(2, len(df) + 2):
                    cell = ws.cell(row=r, column=i)
                    if cell.value:
                        cell.hyperlink = str(cell.value)
                        cell.style = "Hyperlink"


# --------------------------------------------------------------------------- self-check

def self_check():
    """Plain asserts, zero network.  Fails loudly if anyone reverts to a rank compare."""
    U = "u"
    G1, G2 = {"g1"}, {"g1", "g2"}
    T = {"g1": "Alpha", "g2": "Beta"}

    def c(la, ma, lg=None, mg=None, lown=U, mown=U):
        return covers(la, lg, lown, ma, mg, mown, T)[0]

    # 1. all 18 measured matrix rows.  (la, ma, count, expected covers())
    #    'MISSING' rows are decided by state, not covers, and are asserted via verdict below.
    matrix = [
        ("public", "public", 146, True), ("public", "private", 95, True),
        ("shared", "shared", 67, True), ("MISSING", "private", 22, None),
        ("private", "private", 22, True), ("org", "org", 20, True),
        ("shared", "private", 17, True), ("org", "private", 10, True),
        ("MISSING", "shared", 10, None), ("private", "shared", 7, False),
        ("public", "org", 6, True), ("MISSING", "public", 5, None),
        ("org", "public", 3, False), ("private", "org", 2, False),
        ("MISSING", "org", 2, None), ("shared", "org", 1, False),
        ("private", "public", 1, False), ("org", "shared", 1, True),
    ]
    assert len(matrix) == 18
    broken = dead = 0
    for la, ma, cnt, want in matrix:
        state = "MISSING_ITEM" if la == "MISSING" else "item"
        lg = G1 if la == "shared" else None
        mg = G1 if ma == "shared" else set()
        if want is not None:
            assert c(la, ma, lg, mg) is want, "matrix row %s/%s" % (ma, la)
        v = verdict(state, "", la, lg, U, ma, mg, U, T)[0]
        if v == "BROKEN":
            broken += cnt
        elif v == "DEAD":
            dead += cnt
    # shared/shared x67 is group-dependent and scores OK with equal groups, so it lands in
    # neither weighted bucket -- exactly as the blueprint excludes it.
    # 2. weighted totals.  13 BROKEN under the old rank model; the +1 is org-map/group-layer,
    #    deliberately reclassified -- 'org' and 'shared' are incomparable, not equal.
    assert (broken, dead) == (14, 39), (broken, dead)

    # 3. group containment, both directions, and the reason names the missing title
    ok, why = covers("shared", G1, U, "shared", G2, U, T)
    assert ok is False and "Beta" in why, why
    assert covers("shared", G2, U, "shared", G1, U, T)[0] is True

    # 4. a map shared to zero groups normalizes to private.  Probed with an unknown-group
    #    layer: without normalization this is UNKNOWN, with it the map is owner-only -> OK.
    assert verdict("item", "", "shared", None, "other", "shared", set(), U, T)[0] == "OK"
    assert verdict("item", "", "shared", None, "other", "shared", G1, U, T)[0] == "UNKNOWN"

    # 5. private layer owned by someone else is invisible even inside a private map
    assert c("private", "private", lown="bob", mown="alice") is False
    assert c("private", "private", lown="alice", mown="alice") is True

    # 6. unknown group membership is UNKNOWN, never BROKEN
    assert verdict("item", "", "shared", None, U, "shared", G1, U, T)[0] == "UNKNOWN"

    # 7. probe error-code map
    for code in (499, 403, 401, 498):
        assert code_state(code)[0] == "SECURED", code
    assert code_state(404) == ("DEAD", "code 404")
    # 400 "Invalid URL" is how ArcGIS REST reports a service that no longer exists.
    # Verified live: the org_admin session gets the same 400 on all 12 sampled urls.
    assert code_state(400) == ("DEAD", "code 400")
    assert code_state(500)[0] == "UNKNOWN"
    assert verdict("SECURED", "", None, None, None, "public", None, U)[0] == "BROKEN"
    assert verdict("SECURED", "", None, None, None, "org", None, U)[0] == "UNKNOWN"
    assert verdict("PUBLIC", "", "public", set(), "", "public", set(), U)[0] == "OK"

    # 8. circuit breaker returns before any request is attempted
    fails = collections.Counter({"dead.example.gov": MAX_HOST_FAILS})
    cache = {}
    assert probe("https://dead.example.gov/x/FeatureServer/0", cache, fails) == \
        ("UNREACHABLE", "host circuit-open")
    assert cache == {}, "circuit-open must not poison the url cache"

    # 9. walk: GroupLayer children, tables, basemap; shells and junk excluded
    out = []
    walk({"operationalLayers": [
        {"title": "Grp", "layerType": "GroupLayer", "layers": [
            {"title": "Kid", "url": "https://x/FeatureServer/0", "itemId": "k1"}]},
        {"title": "Solo", "url": "https://x/FeatureServer/1"},
        {"title": "Svc", "url": "https://x/MapServer", "layers": [{"id": 0, "popupInfo": {}}]}],
        "tables": [{"title": "Tbl", "itemId": "t1"}],
        "baseMap": {"baseMapLayers": [{"title": "BM", "url": "https://b/MapServer"}]}}, out)
    got = {(s, l.get("title")) for s, l in out}
    assert ("operationalLayers", "Kid") in got and ("operationalLayers", "Grp") not in got
    assert ("tables", "Tbl") in got and ("baseMap", "BM") in got
    assert len(out) == 5, out                # Kid, Solo, Svc, Tbl, BM -- sublayer override dropped
    for junk in (None, {}, [], "x"):
        o = []
        walk(junk, o)
        assert o == []

    # 10. iso takes MILLISECONDS
    assert iso(0) == "" and iso(None) == ""
    assert iso(1700000000000).startswith("2023-11-"), iso(1700000000000)

    # 11. severity band sorts as text
    assert sorted(["2 Medium", "1 High", "4 Info", "3 Low"]) == \
        ["1 High", "2 Medium", "3 Low", "4 Info"]

    # 12. strict feature-layer regex: the service root cannot be counted
    assert FS_RE.search("https://x/arcgis/rest/services/a/featureserver/0")
    assert not FS_RE.search("https://x/arcgis/rest/services/a/featureserver")
    assert not FS_RE.search("https://x/arcgis/rest/services/a/mapserver/0")

    # 13. a hostile item title must not survive as a live Excel formula
    import openpyxl
    _wb = openpyxl.Workbook()
    _ws = _wb.active
    _ws["A1"] = '=cmd|" /C calc"!A0'
    assert _ws["A1"].data_type == "f", "openpyxl no longer auto-types formulas; revisit defuse()"
    defuse(_ws)
    assert _ws["A1"].data_type == "s"
    assert _ws["A1"].value == '=cmd|" /C calc"!A0', "defuse must not alter the text"

    # 14. unknown MAP groups are UNKNOWN, never a silent OK (mirrors the layer-side rule)
    assert covers("shared", G1, U, "shared", None, U, T)[0] is None
    assert verdict("item", "", "shared", G1, U, "shared", None, U, T)[0] == "UNKNOWN"

    # 15. an org with no web maps must summarize, not crash
    assert len(summarize([], {"maps scanned": 0})) > 0

    # 16. a corrupt nested map shape raises out of walk so collect() can record it as an
    #     error row, rather than being silently dropped
    for corrupt in ({"baseMap": "corrupt"}, {"tables": 5},
                    {"operationalLayers": [{"url": "x", "layers": 7}]}):
        try:
            walk(corrupt, [])
        except (AttributeError, TypeError):
            pass
        else:
            raise AssertionError("walk silently accepted %r" % (corrupt,))

    # 17. control characters are stripped rather than crashing the write
    assert scrub(pd.DataFrame({"a": ["ok\x02bad"]}))["a"][0] == "okbad"

    # 18. a DEAD reason must carry the real code, not a hardcoded 404
    assert "400" in verdict("DEAD", "code 400", None, None, None, "public", None, U)[1]

    print("self-check passed")


# --------------------------------------------------------------------------- main

def main():
    p = argparse.ArgumentParser(description="Audit AGOL web maps for layers their viewers cannot see.")
    default = os.path.join(OUTPUT_DIR, "%s_%s.xlsx"
                           % (OUTPUT_PREFIX, datetime.datetime.now().strftime("%Y%m%d_%H%M%S")))
    p.add_argument("-o", "--out", default=default, help="output .xlsx (default: %(default)s)")
    p.add_argument("--limit", "--max", type=int, dest="limit", metavar="N", default=DEFAULT_LIMIT,
                   help="cap maps scanned; smoke-test before the full run")
    p.add_argument("--deep", action="store_true", default=DEEP_DEFAULT,
                   help="add FeatureCount + LastEditUTC (~30s)")
    p.add_argument("--self-check", action="store_true", help="run asserts, no network, and exit")
    a = p.parse_args()
    if a.self_check:
        self_check()
        return 0

    t0 = time.time()
    gis = connect()
    org = gis.properties["name"]
    org_id = gis.properties["id"]
    print("Connected to %s as %s" % (org, gis.users.me.username))

    # `a.limit is None` rather than a truthiness test: --limit 0 must mean zero, not "all".
    maps = gis.content.search("", item_type="Web Map",
                              max_items=MAX_MAPS if a.limit is None else a.limit)
    print("Scanning %d web maps..." % len(maps))
    refs, errors = collect(gis, maps)

    ids = {l.get("itemId") for _, _, l in refs if l.get("itemId")}
    print("%d layer refs, %d distinct layer items" % (len(refs), len(ids)))
    items, unresolved = resolve_items(gis, ids)
    gids, gtitle = fetch_groups(gis, ids | {m["id"] for m in maps})

    probes, fails = {}, collections.Counter()
    rows = build(refs, items, gids, gtitle, probes, fails, org_id, unresolved) + error_rows(errors)
    df = pd.DataFrame(rows, columns=ALL_COLS).fillna("")

    if a.deep:
        stats = deep_stats(gis, set(df.LayerUrl) - {""})
        df["FeatureCount"] = [stats.get(u, ("", ""))[0] for u in df.LayerUrl]
        df["LastEditUTC"] = [stats.get(u, ("", ""))[1] for u in df.LayerUrl]

    meta = {
        "organization": org, "signed in as": gis.users.me.username,
        "started": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "maps scanned": len(maps), "maps that failed to read": len(errors),
        "layer refs": len(refs), "distinct layer items": len(ids),
        "distinct probed urls": len(probes),
        "feature counts": "included (--deep)" if a.deep else "skipped -- re-run with --deep",
    }
    findings = df[df.Verdict != "OK"].sort_values(
        ["Severity", "MapViews", "MapTitle", "LayerTitle"],
        ascending=[True, False, True, True])[FIND_COLS]
    meta["elapsed s"] = round(time.time() - t0, 1)
    written = save_report(a.out, findings, df, summarize(rows, meta))

    print("\n%d findings of %d layer refs in %.0fs -> %s"
          % (len(findings), len(df), time.time() - t0, written))
    for v, n in df.Verdict.value_counts().items():
        print("  %-8s %d" % (v, n))
    return 0


if __name__ == "__main__":
    sys.exit(main())
